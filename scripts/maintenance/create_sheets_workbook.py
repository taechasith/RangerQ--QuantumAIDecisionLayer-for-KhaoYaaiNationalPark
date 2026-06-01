from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TABLES = {
    "User": ["id", "email", "name", "passwordHash", "role", "createdAt", "updatedAt"],
    "Park": ["id", "name", "country", "timezone", "boundaryGeojson", "notes", "createdAt", "updatedAt"],
    "Zone": [
        "id",
        "parkId",
        "code",
        "name",
        "zoneType",
        "geometryGeojson",
        "centroidLat",
        "centroidLng",
        "areaHectares",
        "baseFireRisk",
        "baseWildlifeRisk",
        "accessDifficulty",
        "rangerStationDistanceKm",
        "isSynthetic",
        "notes",
        "createdAt",
        "updatedAt",
    ],
    "DataSource": ["id", "parkId", "type", "name", "status", "lastSyncedAt", "freshnessWarning", "config", "createdAt", "updatedAt"],
    "ImportBatch": ["id", "dataSourceId", "sourceType", "filename", "status", "recordCount", "errorMessage", "metadata", "createdAt"],
    "Observation": ["id", "importBatchId", "zoneId", "externalId", "timestamp", "lat", "lng", "eventType", "species", "threatType", "patrolTeam", "confidence", "notes", "raw", "createdAt"],
    "FireHotspot": ["id", "zoneId", "source", "satellite", "timestamp", "lat", "lng", "brightness", "confidence", "frp", "raw", "createdAt"],
    "WeatherSnapshot": ["id", "zoneId", "timestamp", "temperature2m", "relativeHumidity2m", "windSpeed10m", "windDirection10m", "windGusts10m", "precipitation", "soilMoisture0To1cm", "soilMoisture1To3cm", "raw", "createdAt"],
    "CameraDetection": ["id", "importBatchId", "zoneId", "cameraId", "timestamp", "lat", "lng", "species", "count", "confidence", "imageUrl", "notes", "raw", "createdAt"],
    "VisitorPressure": ["id", "importBatchId", "zoneId", "date", "zoneCode", "visitorCount", "vehicleCount", "busCount", "source", "notes", "raw", "createdAt"],
    "ManualNote": ["id", "zoneId", "author", "timestamp", "category", "severity", "lat", "lng", "note", "createdAt"],
    "RiskRun": ["id", "status", "weights", "version", "startedAt", "completedAt", "errorMessage", "createdAt"],
    "ZoneRiskScore": ["id", "riskRunId", "zoneId", "fireRisk", "wildlifeRisk", "combinedPriority", "operationalFeasibility", "label", "topFactors", "recommendedAction", "freshnessWarning", "createdAt"],
    "OptimizationRun": ["id", "riskRunId", "method", "status", "maxZones", "maxPatrolHours", "rangerTeams", "requireFireCoverage", "requireWildlifeCoverage", "coverageScore", "travelPenalty", "estimatedPatrolHours", "quboPayload", "qbraidJobId", "qbraidStatus", "qbraidResult", "fallbackReason", "createdAt", "completedAt"],
    "SelectedZone": ["id", "optimizationRunId", "zoneId", "rank", "reason", "estimatedHours", "scoreContribution", "createdAt"],
    "PatrolPlan": ["id", "optimizationRunId", "teamName", "route", "objective", "why", "estimatedHours", "safetyNote", "fallback", "createdAt"],
    "AuditLog": ["id", "userId", "action", "entity", "entityId", "metadata", "createdAt"],
}

ZONES = [
    ("zone-ky-bnd-01", "khao-yai-national-park", "KY-BND-01", "Northwest Boundary Corridor", "BOUNDARY", 14.52, 101.32, 420, 72, 58, 46, 7.4),
    ("zone-ky-bnd-02", "khao-yai-national-park", "KY-BND-02", "Pak Chong Village Edge", "VILLAGE_EDGE", 14.48, 101.37, 310, 84, 76, 34, 4.2),
    ("zone-ky-for-03", "khao-yai-national-park", "KY-FOR-03", "Dry Dipterocarp Ridge", "FOREST", 14.44, 101.35, 680, 79, 45, 68, 9.1),
    ("zone-ky-trl-04", "khao-yai-national-park", "KY-TRL-04", "Haew Suwat Trail Sector", "TRAIL", 14.43, 101.42, 260, 54, 64, 38, 5.6),
    ("zone-ky-vis-05", "khao-yai-national-park", "KY-VIS-05", "Visitor Center Pressure Zone", "VISITOR_AREA", 14.44, 101.38, 150, 62, 52, 20, 1.1),
    ("zone-ky-rd-06", "khao-yai-national-park", "KY-RD-06", "Thanarat Road Wildlife Crossing", "ROAD", 14.41, 101.39, 190, 48, 82, 28, 3.5),
    ("zone-ky-wat-07", "khao-yai-national-park", "KY-WAT-07", "Lam Takhong Water Source", "WATER_SOURCE", 14.39, 101.44, 230, 35, 78, 42, 6.8),
    ("zone-ky-for-08", "khao-yai-national-park", "KY-FOR-08", "Central Evergreen Interior", "FOREST", 14.35, 101.45, 940, 44, 69, 75, 11.2),
    ("zone-ky-bnd-09", "khao-yai-national-park", "KY-BND-09", "Southern Farm Interface", "VILLAGE_EDGE", 14.28, 101.43, 360, 81, 88, 52, 8.9),
    ("zone-ky-rs-10", "khao-yai-national-park", "KY-RS-10", "Nang Rong Ranger Station", "RANGER_STATION", 14.31, 101.37, 80, 37, 40, 16, 0),
    ("zone-ky-bnd-11", "khao-yai-national-park", "KY-BND-11", "Eastern Boundary Firebreak", "BOUNDARY", 14.37, 101.56, 410, 86, 55, 57, 10.7),
    ("zone-ky-trl-12", "khao-yai-national-park", "KY-TRL-12", "Khao Khieo Trail Approach", "TRAIL", 14.46, 101.49, 280, 51, 61, 64, 7.9),
    ("zone-ky-for-13", "khao-yai-national-park", "KY-FOR-13", "Northeast Forest Block", "FOREST", 14.55, 101.51, 760, 66, 57, 71, 12.1),
    ("zone-ky-wat-14", "khao-yai-national-park", "KY-WAT-14", "Mo Singto Wetland Fringe", "WATER_SOURCE", 14.47, 101.44, 210, 29, 84, 39, 5.3),
    ("zone-ky-rd-15", "khao-yai-national-park", "KY-RD-15", "Night Crossing Road Bend", "ROAD", 14.34, 101.49, 170, 43, 90, 31, 6.6),
    ("zone-ky-bnd-16", "khao-yai-national-park", "KY-BND-16", "Western Smoke Report Zone", "BOUNDARY", 14.38, 101.27, 500, 91, 49, 62, 10.2),
    ("zone-ky-vis-17", "khao-yai-national-park", "KY-VIS-17", "Campground Visitor Cluster", "VISITOR_AREA", 14.42, 101.46, 120, 68, 63, 22, 2.8),
    ("zone-ky-for-18", "khao-yai-national-park", "KY-FOR-18", "Remote Southeast Forest", "FOREST", 14.21, 101.53, 820, 59, 70, 82, 14.4),
    ("zone-ky-bnd-19", "khao-yai-national-park", "KY-BND-19", "Elephant Conflict Boundary", "VILLAGE_EDGE", 14.25, 101.34, 390, 73, 94, 48, 8.1),
    ("zone-ky-rs-20", "khao-yai-national-park", "KY-RS-20", "Northern Patrol Base", "RANGER_STATION", 14.5, 101.41, 95, 41, 46, 18, 0),
]


def point_geometry(lat, lng):
    return (
        '{"type":"Feature","properties":{"synthetic":true,"warning":"Demo point geometry. '
        'Replace with official Khao Yai GIS polygons before operational use."},'
        f'"geometry":{{"type":"Point","coordinates":[{lng},{lat}]}}}}'
    )


def style_sheet(sheet):
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22


def main():
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name, headers in TABLES.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        style_sheet(sheet)

    now = "2026-05-31T00:00:00+07:00"
    workbook["Park"].append([
        "khao-yai-national-park",
        "Khao Yai National Park",
        "Thailand",
        "Asia/Bangkok",
        '{"type":"FeatureCollection","features":[],"synthetic":true}',
        "Operational demo seeded with synthetic zones. Replace with official GIS and field data before real use.",
        now,
        now,
    ])

    workbook["User"].append([
        "admin-user",
        "admin@rangerq.local",
        "RangerQ Admin",
        "managed-in-app-or-apps-script",
        "ADMIN",
        now,
        now,
    ])

    for zone in ZONES:
        zone_id, park_id, code, name, zone_type, lat, lng, area, fire, wildlife, access, station_distance = zone
        workbook["Zone"].append([
            zone_id,
            park_id,
            code,
            name,
            zone_type,
            point_geometry(lat, lng),
            lat,
            lng,
            area,
            fire,
            wildlife,
            access,
            station_distance,
            "TRUE",
            "Synthetic demo zone. Replace with official Khao Yai GIS polygon.",
            now,
            now,
        ])

    for source_id, source_type, name, status in [
        ("source-smart", "SMART_EXPORT", "SMART patrol export", "demo"),
        ("source-camera", "CAMERA_AI", "Camera and AI detections", "demo"),
        ("source-visitors", "VISITOR_CSV", "Visitor pressure CSV", "demo"),
        ("source-weather", "OPEN_METEO", "Open-Meteo weather", "configured"),
        ("source-gbif", "GBIF_OCCURRENCE", "GBIF species occurrences", "configured"),
    ]:
        workbook["DataSource"].append([
            source_id,
            "khao-yai-national-park",
            source_type,
            name,
            status,
            "",
            "Demo data source until official data is imported." if status == "demo" else "",
            "{}",
            now,
            now,
        ])

    workbook["AuditLog"].append([
        "audit-seed-demo-data",
        "",
        "SEED_DEMO_DATA",
        "Park",
        "khao-yai-national-park",
        '{"zones":20,"synthetic":true}',
        now,
    ])

    output_dir = Path(".rangerq")
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "RangerQ_Database.xlsx"
    workbook.save(output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()
